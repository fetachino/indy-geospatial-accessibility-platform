"""Download, cache, checksum, and validate catalog datasets."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import shutil
import tempfile
import zipfile
from collections.abc import Callable, Iterable, Iterator, Mapping, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from indy_accessibility_data.catalog import Dataset, ValidationRules

USER_AGENT = (
    "indy-geospatial-accessibility-platform/0.1 "
    "(+https://github.com/fetachino/indy-geospatial-accessibility-platform)"
)
DEFAULT_CACHE_DIR = Path("data/raw")
DEFAULT_TIMEOUT_SECONDS = 90.0


class AcquisitionError(RuntimeError):
    """Raised when a dataset cannot be acquired safely."""


class SourceValidationError(AcquisitionError):
    """Raised when a downloaded or existing source fails validation."""


@dataclass(frozen=True)
class AcquisitionResult:
    """Validated cache result suitable for CLI and tests."""

    dataset_id: str
    path: Path
    sha256: str
    bytes: int
    downloaded: bool
    manifest_path: Path


ResponseOpener = Callable[[Request, float], Any]


def acquire_dataset(
    dataset: Dataset,
    *,
    cache_dir: Path = DEFAULT_CACHE_DIR,
    force: bool = False,
    validate_existing: bool = False,
    timeout: float = DEFAULT_TIMEOUT_SECONDS,
    environment: Mapping[str, str] | None = None,
    opener: ResponseOpener | None = None,
) -> AcquisitionResult:
    """Acquire one dataset atomically and validate it before caching."""
    environment = os.environ if environment is None else environment
    target_dir = cache_dir / dataset.id
    target = target_dir / dataset.cache_filename
    manifest = target.with_suffix(f"{target.suffix}.metadata.json")

    if validate_existing:
        if not target.is_file():
            raise AcquisitionError(
                f"{dataset.id}: no cached file exists at {target}; "
                "follow the catalog manual_fallback first"
            )
        validate_source(target, dataset)
        return _result(dataset.id, target, manifest, downloaded=False)

    if target.is_file() and not force:
        validate_source(target, dataset)
        return _result(dataset.id, target, manifest, downloaded=False)

    url = _render_url(dataset, environment)
    target_dir.mkdir(parents=True, exist_ok=True)
    opener = _default_opener if opener is None else opener

    request = Request(url, headers={"User-Agent": USER_AGENT})
    temporary_path: Path | None = None
    response_headers: Mapping[str, str] = {}
    try:
        with opener(request, timeout) as response:
            status = getattr(response, "status", 200)
            if status < 200 or status >= 300:
                raise AcquisitionError(
                    f"{dataset.id}: source returned unexpected HTTP status {status}"
                )
            response_headers = dict(getattr(response, "headers", {}))
            content_type = _header_value(response_headers, "Content-Type")
            _validate_content_type(dataset, content_type)
            with tempfile.NamedTemporaryFile(
                mode="wb", dir=target_dir, prefix=f".{target.name}.", delete=False
            ) as temporary:
                temporary_path = Path(temporary.name)
                shutil.copyfileobj(response, temporary)

        validate_source(temporary_path, dataset)
        temporary_path.replace(target)
        _write_manifest(dataset, target, manifest, response_headers)
        temporary_path = None
    except (HTTPError, URLError, TimeoutError, OSError) as error:
        raise AcquisitionError(
            f"{dataset.id}: unable to download source: {_safe_network_error(error)}; "
            f"manual fallback: {dataset.manual_fallback}"
        ) from error
    finally:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)

    return _result(dataset.id, target, manifest, downloaded=True)


def validate_source(path: Path, dataset: Dataset) -> None:
    """Apply the catalog's format and minimum-schema checks."""
    if not path.is_file():
        raise SourceValidationError(f"{dataset.id}: source file does not exist: {path}")
    rules = dataset.validation
    validators: dict[str, Callable[[Path, Dataset, ValidationRules], None]] = {
        "census_json": _validate_census_json,
        "csv_zip": _validate_csv_zip,
        "geojson": _validate_geojson,
        "gtfs_zip": _validate_gtfs_zip,
        "shapefile_zip": _validate_shapefile_zip,
        "xlsx": _validate_xlsx,
    }
    try:
        validators[rules.kind](path, dataset, rules)
    except SourceValidationError:
        raise
    except (csv.Error, json.JSONDecodeError, KeyError, TypeError, ValueError) as error:
        raise SourceValidationError(
            f"{dataset.id}: {rules.kind} validation failed: {error}"
        ) from error
    except (InvalidFileException, OSError, zipfile.BadZipFile) as error:
        raise SourceValidationError(
            f"{dataset.id}: expected {dataset.expected_format}, "
            f"but the file is invalid: {error}"
        ) from error


def _render_url(dataset: Dataset, environment: Mapping[str, str]) -> str:
    url = dataset.source_url
    missing = [
        name for name in dataset.requires_environment if not environment.get(name)
    ]
    if missing:
        names = ", ".join(missing)
        raise AcquisitionError(
            f"{dataset.id}: required environment variable(s) not set: {names}; "
            f"manual fallback: {dataset.manual_fallback}"
        )
    for name in dataset.requires_environment:
        url = url.replace(f"{{{name}}}", environment[name])
    return url


def _default_opener(request: Request, timeout: float) -> Any:
    return urlopen(request, timeout=timeout)


def _validate_content_type(dataset: Dataset, content_type: str) -> None:
    if not content_type:
        return
    media_type = content_type.split(";", maxsplit=1)[0].strip().lower()
    allowed = {
        "census_json": {"application/json", "text/json"},
        "csv_zip": {"application/zip", "application/octet-stream"},
        "geojson": {
            "application/geo+json",
            "application/json",
            "text/plain",
        },
        "gtfs_zip": {"application/zip", "application/octet-stream"},
        "shapefile_zip": {
            "application/zip",
            "application/octet-stream",
            "application/x-zip-compressed",
        },
        "xlsx": {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/octet-stream",
        },
    }
    if media_type not in allowed[dataset.validation.kind]:
        raise SourceValidationError(
            f"{dataset.id}: expected {dataset.expected_format}, "
            f"but server returned Content-Type {content_type!r}"
        )


def _validate_geojson(path: Path, dataset: Dataset, rules: ValidationRules) -> None:
    document = json.loads(path.read_text(encoding="utf-8-sig"))
    if document.get("type") != "FeatureCollection":
        raise SourceValidationError(
            f"{dataset.id}: expected a GeoJSON FeatureCollection"
        )
    features = document.get("features")
    if not isinstance(features, list):
        raise SourceValidationError(f"{dataset.id}: GeoJSON features must be a list")
    _require_minimum_records(dataset, len(features), rules.minimum_records)
    required_fields = _string_list(rules.required_fields, dataset)
    allowed_geometry_types = set(rules.geometry_types or [])
    observed_coordinate = False
    for index, feature in enumerate(features):
        properties = feature.get("properties") or {}
        _require_fields(dataset, properties.keys(), required_fields, f"feature {index}")
        _require_field_values(dataset, properties, rules.field_equals)
        geometry = feature.get("geometry") or {}
        geometry_type = geometry.get("type")
        if geometry_type not in allowed_geometry_types:
            raise SourceValidationError(
                f"{dataset.id}: feature {index} has unexpected geometry type "
                f"{geometry_type!r}"
            )
        for longitude, latitude in _coordinate_pairs(geometry.get("coordinates")):
            observed_coordinate = True
            _require_bounds(dataset, longitude, latitude, rules.bounds)
    if not observed_coordinate:
        raise SourceValidationError(f"{dataset.id}: no usable coordinates found")


def _validate_gtfs_zip(path: Path, dataset: Dataset, rules: ValidationRules) -> None:
    with zipfile.ZipFile(path) as archive:
        names = {_normalized_member(name): name for name in archive.namelist()}
        _require_members(dataset, names.keys(), rules.required_members or [])
        required_fields = rules.required_fields
        if not isinstance(required_fields, dict):
            raise SourceValidationError(f"{dataset.id}: GTFS rules require a field map")
        for filename, expected_fields in required_fields.items():
            member = names[_normalized_member(filename)]
            header, rows = _csv_header_and_rows(archive, member)
            _require_fields(dataset, header, expected_fields, filename)
            _require_minimum_records(dataset, len(rows), rules.minimum_records)

        stops_member = names[_normalized_member("stops.txt")]
        with archive.open(stops_member) as raw:
            reader = csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8-sig"))
            observed_in_bounds = False
            for row in reader:
                try:
                    longitude = float(row["stop_lon"])
                    latitude = float(row["stop_lat"])
                except (KeyError, TypeError, ValueError) as error:
                    raise SourceValidationError(
                        f"{dataset.id}: stops.txt has invalid stop coordinates"
                    ) from error
                _require_world_coordinate(dataset, longitude, latitude)
                if _is_in_bounds(longitude, latitude, rules.bounds):
                    observed_in_bounds = True
            if rules.bounds and not observed_in_bounds:
                raise SourceValidationError(
                    f"{dataset.id}: no GTFS stops fall within the expected "
                    "coverage bounds"
                )


def _validate_shapefile_zip(
    path: Path, dataset: Dataset, rules: ValidationRules
) -> None:
    _require_minimum_bytes(dataset, path, rules.minimum_bytes)
    with zipfile.ZipFile(path) as archive:
        names = {_normalized_member(name) for name in archive.namelist()}
        _require_members(dataset, names, rules.required_members or [])


def _validate_csv_zip(path: Path, dataset: Dataset, rules: ValidationRules) -> None:
    _require_minimum_bytes(dataset, path, rules.minimum_bytes)
    with zipfile.ZipFile(path) as archive:
        csv_members = [
            name for name in archive.namelist() if name.lower().endswith(".csv")
        ]
        if not csv_members:
            raise SourceValidationError(f"{dataset.id}: ZIP contains no CSV file")
        header, rows = _csv_header_and_rows(archive, csv_members[0])
        _require_fields(
            dataset, header, _string_list(rules.required_fields, dataset), "CSV"
        )
        _require_minimum_records(dataset, len(rows), rules.minimum_records)


def _validate_xlsx(path: Path, dataset: Dataset, rules: ValidationRules) -> None:
    if path.read_bytes()[:4] != b"PK\x03\x04":
        raise SourceValidationError(f"{dataset.id}: XLSX is not an OOXML ZIP file")
    with path.open("rb") as source:
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            for sheet_name, expected_fields in (rules.required_sheets or {}).items():
                if sheet_name not in workbook.sheetnames:
                    raise SourceValidationError(
                        f"{dataset.id}: workbook is missing sheet {sheet_name!r}"
                    )
                sheet = workbook[sheet_name]
                values = sheet.iter_rows(values_only=True)
                try:
                    header_row = next(values)
                except StopIteration as error:
                    raise SourceValidationError(
                        f"{dataset.id}: sheet {sheet_name!r} is empty"
                    ) from error
                header = [str(value) for value in header_row if value is not None]
                _require_fields(dataset, header, expected_fields, sheet_name)
                count = sum(
                    1 for row in values if any(value is not None for value in row)
                )
                _require_minimum_records(dataset, count, rules.minimum_records)
        finally:
            workbook.close()


def _validate_census_json(path: Path, dataset: Dataset, rules: ValidationRules) -> None:
    document = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(document, list) or not document:
        raise SourceValidationError(
            f"{dataset.id}: Census response must be a row array"
        )
    header = document[0]
    if not isinstance(header, list) or not all(
        isinstance(value, str) for value in header
    ):
        raise SourceValidationError(
            f"{dataset.id}: Census response has no valid header"
        )
    required_fields = _string_list(rules.required_fields, dataset)
    _require_fields(dataset, header, required_fields, "Census header")
    rows = document[1:]
    _require_minimum_records(dataset, len(rows), rules.minimum_records)
    positions = {name: header.index(name) for name in required_fields}
    for row in rows:
        if not isinstance(row, list) or len(row) != len(header):
            raise SourceValidationError(f"{dataset.id}: Census row width is invalid")
        values = {field: str(row[index]) for field, index in positions.items()}
        _require_field_values(dataset, values, rules.field_equals)


def _write_manifest(
    dataset: Dataset,
    target: Path,
    manifest: Path,
    response_headers: Mapping[str, str],
) -> None:
    metadata = {
        "dataset_id": dataset.id,
        "source_url_template": dataset.source_url,
        "source_organization": dataset.source_organization,
        "retrieved_at": datetime.now(UTC).isoformat(),
        "sha256": _sha256(target),
        "bytes": target.stat().st_size,
        "etag": _header_value(response_headers, "ETag") or None,
        "last_modified": _header_value(response_headers, "Last-Modified") or None,
    }
    temporary = manifest.with_suffix(f"{manifest.suffix}.tmp")
    temporary.write_text(json.dumps(metadata, indent=2) + "\n", encoding="utf-8")
    temporary.replace(manifest)


def _result(
    dataset_id: str, target: Path, manifest: Path, *, downloaded: bool
) -> AcquisitionResult:
    return AcquisitionResult(
        dataset_id=dataset_id,
        path=target,
        sha256=_sha256(target),
        bytes=target.stat().st_size,
        downloaded=downloaded,
        manifest_path=manifest,
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _csv_header_and_rows(
    archive: zipfile.ZipFile, member: str
) -> tuple[list[str], list[list[str]]]:
    with archive.open(member) as raw:
        reader = csv.reader(io.TextIOWrapper(raw, encoding="utf-8-sig"))
        try:
            header = next(reader)
        except StopIteration:
            return [], []
        return header, list(reader)


def _normalized_member(name: str) -> str:
    return Path(name.replace("\\", "/")).name.casefold()


def _require_members(
    dataset: Dataset, actual: Iterable[str], expected: Sequence[str]
) -> None:
    normalized = {_normalized_member(name) for name in actual}
    missing = [name for name in expected if _normalized_member(name) not in normalized]
    if missing:
        raise SourceValidationError(
            f"{dataset.id}: archive is missing required files: {', '.join(missing)}"
        )


def _require_fields(
    dataset: Dataset,
    actual: Sequence[str] | Any,
    expected: Sequence[str],
    location: str,
) -> None:
    available = {str(value) for value in actual}
    missing = [field for field in expected if field not in available]
    if missing:
        raise SourceValidationError(
            f"{dataset.id}: {location} is missing required fields: {', '.join(missing)}"
        )


def _require_field_values(
    dataset: Dataset,
    values: Mapping[str, Any],
    expected: Mapping[str, str] | None,
) -> None:
    for field, required_value in (expected or {}).items():
        if str(values.get(field, "")) != required_value:
            raise SourceValidationError(
                f"{dataset.id}: field {field!r} must equal {required_value!r}"
            )


def _require_minimum_records(
    dataset: Dataset, actual: int, minimum: int | None
) -> None:
    if minimum is not None and actual < minimum:
        raise SourceValidationError(
            f"{dataset.id}: expected at least {minimum} records, found {actual}"
        )


def _require_minimum_bytes(dataset: Dataset, path: Path, minimum: int | None) -> None:
    if minimum is not None and path.stat().st_size < minimum:
        raise SourceValidationError(
            f"{dataset.id}: expected at least {minimum} bytes, "
            f"found {path.stat().st_size}"
        )


def _coordinate_pairs(coordinates: Any) -> Iterator[tuple[float, float]]:
    if (
        isinstance(coordinates, list)
        and len(coordinates) >= 2
        and all(isinstance(value, int | float) for value in coordinates[:2])
    ):
        yield float(coordinates[0]), float(coordinates[1])
        return
    if isinstance(coordinates, list):
        for nested in coordinates:
            yield from _coordinate_pairs(nested)


def _require_bounds(
    dataset: Dataset,
    longitude: float,
    latitude: float,
    bounds: tuple[float, float, float, float] | None,
) -> None:
    _require_world_coordinate(dataset, longitude, latitude)
    if bounds and not _is_in_bounds(longitude, latitude, bounds):
        raise SourceValidationError(
            f"{dataset.id}: coordinate ({longitude}, {latitude}) is outside "
            "expected coverage bounds"
        )


def _require_world_coordinate(
    dataset: Dataset, longitude: float, latitude: float
) -> None:
    if not -180 <= longitude <= 180 or not -90 <= latitude <= 90:
        raise SourceValidationError(
            f"{dataset.id}: invalid longitude/latitude ({longitude}, {latitude})"
        )


def _is_in_bounds(
    longitude: float,
    latitude: float,
    bounds: tuple[float, float, float, float] | None,
) -> bool:
    if bounds is None:
        return True
    west, south, east, north = bounds
    return west <= longitude <= east and south <= latitude <= north


def _string_list(
    value: list[str] | dict[str, list[str]] | None, dataset: Dataset
) -> list[str]:
    if not isinstance(value, list):
        raise SourceValidationError(
            f"{dataset.id}: validation rules require a field list"
        )
    return value


def _header_value(headers: Mapping[str, str], name: str) -> str:
    for key, value in headers.items():
        if key.casefold() == name.casefold():
            return value
    return ""


def _safe_network_error(error: BaseException) -> str:
    if isinstance(error, HTTPError):
        return f"HTTP {error.code} {error.reason}"
    if isinstance(error, URLError):
        return str(error.reason)
    return str(error)
