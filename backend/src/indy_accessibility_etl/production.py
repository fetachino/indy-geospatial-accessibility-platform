"""Load the Milestone 1 raw cache into normalized production records.

The raw cache is deliberately discovered at runtime and remains ignored by Git.
This module is also useful in dry-run mode, where it validates and counts the
records without requiring a database connection.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from shapefile import Reader  # type: ignore[import-untyped]
from shapely.geometry import MultiPolygon, shape

from .geometry import (
    GeometryDecision,
    feature_source_id,
    project_geometry,
    validate_geometry,
)
from .migrations import apply_migrations

RAW_ROOT = Path("data/raw")


@dataclass(frozen=True)
class ProductionRecord:
    dataset_id: str
    source_id: str
    name: str
    service_type: str | None
    geometry: Any
    source_crs: str
    address: str | None = None
    attributes: dict[str, Any] | None = None


@dataclass(frozen=True)
class ProductionRun:
    records: tuple[ProductionRecord, ...]
    audit: tuple[dict[str, str], ...]
    unsupported: tuple[str, ...]


DATASET_FILES = {
    "marion_county_boundary": (
        "marion_county_boundary",
        "marion_county_boundary.geojson",
    ),
    "census_block_groups_2024": ("census_block_groups_2024", "tl_2024_18_bg.zip"),
    "indygo_gtfs": ("indygo_gtfs", "indygo_gtfs.zip"),
    "hospital_locations_2023": (
        "hospital_locations_2023",
        "hospital_locations_2023_marion.geojson",
    ),
    "snap_grocery_retailers_2005_2025": (
        "snap_grocery_retailers_2005_2025",
        "snap_retailers_2005_2025.zip",
    ),
    "indiana_school_directory_2025_2026": (
        "indiana_school_directory_2025_2026",
        "indiana_school_directory_2025_2026.xlsx",
    ),
    "indiana_library_locations_2025": (
        "indiana_library_locations_2025",
        "indiana_library_locations_2025.geojson",
    ),
    "ifd_fire_stations": ("ifd_fire_stations", "ifd_fire_stations.geojson"),
}


def discover_cached_sources(raw_root: Path = RAW_ROOT) -> dict[str, Path]:
    """Return only known, present cache files; never search outside the cache."""
    found: dict[str, Path] = {}
    for dataset_id, (directory, filename) in DATASET_FILES.items():
        path = raw_root / directory / filename
        if path.is_file() and path.stat().st_size > 0:
            found[dataset_id] = path
    return found


def _geojson_records(
    dataset_id: str, path: Path, service_type: str | None
) -> list[ProductionRecord]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    features = payload.get("features", [])
    records: list[ProductionRecord] = []
    for index, feature in enumerate(features):
        properties = feature.get("properties") or {}
        geometry = shape(feature["geometry"])
        source_id = feature_source_id(properties, f"{dataset_id}-{index}")
        name = str(
            properties.get("name")
            or properties.get("STATION")
            or properties.get("user_branch_name")
            or source_id
        )
        address = (
            properties.get("address")
            or properties.get("ADDRESS")
            or properties.get("fulladdress")
        )
        records.append(
            ProductionRecord(
                dataset_id,
                source_id,
                name,
                service_type,
                geometry,
                "EPSG:4326",
                address,
                properties,
            )
        )
    return records


def _gtfs_records(path: Path) -> list[ProductionRecord]:
    records: list[ProductionRecord] = []
    with zipfile.ZipFile(path) as archive, archive.open("stops.txt") as stream:
        for row in csv.DictReader(io.TextIOWrapper(stream, encoding="utf-8-sig")):
            records.append(
                ProductionRecord(
                    "indygo_gtfs",
                    row["stop_id"],
                    row["stop_name"],
                    None,
                    shape(
                        {
                            "type": "Point",
                            "coordinates": [
                                float(row["stop_lon"]),
                                float(row["stop_lat"]),
                            ],
                        }
                    ),
                    "EPSG:4326",
                    row.get("stop_desc"),
                    row,
                )
            )
    return records


def _grocery_records(path: Path) -> list[ProductionRecord]:
    records: list[ProductionRecord] = []
    with zipfile.ZipFile(path) as archive:
        csv_name = next(
            name for name in archive.namelist() if name.lower().endswith(".csv")
        )
        with archive.open(csv_name) as stream:
            for row in csv.DictReader(io.TextIOWrapper(stream, encoding="utf-8-sig")):
                if (
                    row.get("State", "").upper() != "IN"
                    or row.get("County", "").upper() not in {"MARION", "MARION COUNTY"}
                    or not row.get("Latitude")
                    or not row.get("Longitude")
                ):
                    continue
                records.append(
                    ProductionRecord(
                        "snap_grocery_retailers_2005_2025",
                        row.get("Record ID", ""),
                        row.get("Store Name", ""),
                        "grocery_store",
                        shape(
                            {
                                "type": "Point",
                                "coordinates": [
                                    float(row["Longitude"]),
                                    float(row["Latitude"]),
                                ],
                            }
                        ),
                        "EPSG:4326",
                        row.get("Address"),
                        row,
                    )
                )
    return records


def _block_group_records(path: Path) -> list[ProductionRecord]:
    with zipfile.ZipFile(path) as archive:
        members = {
            Path(name).suffix.lower(): archive.read(name)
            for name in archive.namelist()
            if Path(name).suffix.lower() in {".shp", ".shx", ".dbf", ".prj"}
        }
    reader = Reader(
        shp=io.BytesIO(members[".shp"]),
        shx=io.BytesIO(members[".shx"]),
        dbf=io.BytesIO(members[".dbf"]),
    )
    records: list[ProductionRecord] = []
    for record, geometry in zip(reader.records(), reader.shapes(), strict=True):
        field_names = [field[0] for field in reader.fields[1:]]
        properties = dict(zip(field_names, record, strict=True))
        records.append(
            ProductionRecord(
                "census_block_groups_2024",
                str(properties["GEOID"]),
                str(properties.get("NAMELSAD", properties["GEOID"])),
                None,
                shape(geometry.__geo_interface__),
                "EPSG:4269",
                attributes=properties,
            )
        )
    return records


def _school_check(path: Path) -> None:
    workbook = load_workbook(path, read_only=True)
    required = {"SCHL", "NPSCHL"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"School workbook is missing sheets: {sorted(missing)}")
    # Addresses are retained for a later, separately documented geocoding step.
    workbook.close()


def _validate_records(
    records: Iterable[ProductionRecord], county_geometry: Any | None
) -> tuple[list[ProductionRecord], list[dict[str, str]]]:
    loaded: list[ProductionRecord] = []
    audit: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for record in records:
        key = (record.dataset_id, record.source_id)
        if key in seen:
            audit.append(
                {
                    "dataset_id": record.dataset_id,
                    "source_id": record.source_id,
                    "action": "duplicate",
                    "reason": "duplicate source identifier",
                }
            )
            continue
        seen.add(key)
        expected = (
            "MultiPolygon"
            if record.dataset_id
            in {"marion_county_boundary", "census_block_groups_2024"}
            else "Point"
        )
        decision: GeometryDecision = validate_geometry(
            record.geometry,
            source_epsg=int(record.source_crs.split(":")[-1]),
            expected_type=expected,
            county_geometry=county_geometry
            if record.dataset_id
            not in {"marion_county_boundary", "census_block_groups_2024"}
            else None,
        )
        audit.append(
            {
                "dataset_id": record.dataset_id,
                "source_id": record.source_id,
                "action": decision.action,
                "reason": decision.reason or "",
            }
        )
        if decision.geometry is not None:
            loaded.append(
                ProductionRecord(
                    record.dataset_id,
                    record.source_id,
                    record.name,
                    record.service_type,
                    decision.geometry,
                    record.source_crs,
                    record.address,
                    record.attributes,
                )
            )
    return loaded, audit


def run_production_etl(raw_root: Path = RAW_ROOT) -> ProductionRun:
    sources = discover_cached_sources(raw_root)
    if not sources:
        raise FileNotFoundError(
            "No recognized Milestone 1 files found under data/raw; "
            "run acquisition first"
        )
    boundary: list[ProductionRecord] = []
    records: list[ProductionRecord] = []
    unsupported: list[str] = []
    for dataset_id, path in sources.items():
        if dataset_id == "marion_county_boundary":
            boundary = _geojson_records(dataset_id, path, None)
            records.extend(boundary)
        elif dataset_id == "census_block_groups_2024":
            records.extend(_block_group_records(path))
        elif dataset_id == "indygo_gtfs":
            records.extend(_gtfs_records(path))
        elif dataset_id == "hospital_locations_2023":
            records.extend(_geojson_records(dataset_id, path, "hospital"))
        elif dataset_id == "snap_grocery_retailers_2005_2025":
            records.extend(_grocery_records(path))
        elif dataset_id == "indiana_library_locations_2025":
            records.extend(_geojson_records(dataset_id, path, "library"))
        elif dataset_id == "ifd_fire_stations":
            records.extend(_geojson_records(dataset_id, path, "fire_station"))
        elif dataset_id == "indiana_school_directory_2025_2026":
            _school_check(path)
            unsupported.append(
                "indiana_school_directory_2025_2026: address-only workbook "
                "requires later geocoding"
            )
    county_geometry = project_geometry(boundary[0].geometry, 4326) if boundary else None
    loaded, audit = _validate_records(records, county_geometry)
    return ProductionRun(tuple(loaded), tuple(audit), tuple(unsupported))


def load_production_to_database(url: str, raw_root: Path = RAW_ROOT) -> ProductionRun:
    """Run production discovery/validation and insert supported rows transactionally."""
    run = run_production_etl(raw_root)
    try:
        import psycopg
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Install psycopg to load production data") from exc
    apply_migrations(url)
    load_run_id = str(uuid.uuid4())
    with psycopg.connect(url) as connection, connection.transaction():
        connection.execute(
            "INSERT INTO etl.load_runs(load_run_id, source_dataset_id, status, "
            "transformation_version, row_count, rejected_count) VALUES "
            "(%s, %s, 'running', %s, %s, %s)",
            (
                load_run_id,
                "milestone1_cached_production",
                "milestone-2-v1",
                len(run.records),
                len(run.audit) - len(run.records),
            ),
        )
        for record in run.records:
            geometry = record.geometry
            if (
                record.dataset_id
                in {"marion_county_boundary", "census_block_groups_2024"}
                and geometry.geom_type == "Polygon"
            ):
                geometry = MultiPolygon([geometry])
            if record.dataset_id == "marion_county_boundary":
                connection.execute(
                    "INSERT INTO boundaries.marion_county(source_id, name, source_crs, "
                    "load_run_id, geometry) "
                    "VALUES (%s, %s, %s, %s, ST_SetSRID(ST_GeomFromText(%s), 26916)) "
                    "ON CONFLICT (source_id) DO UPDATE SET name=EXCLUDED.name, "
                    "geometry=EXCLUDED.geometry, load_run_id=EXCLUDED.load_run_id",
                    (
                        record.source_id,
                        record.name,
                        record.source_crs,
                        load_run_id,
                        geometry.wkt,
                    ),
                )
            elif record.dataset_id == "census_block_groups_2024":
                connection.execute(
                    "INSERT INTO demographics.census_block_groups(geoid, name, "
                    "source_crs, "
                    "load_run_id, geometry) "
                    "VALUES (%s, %s, %s, %s, ST_SetSRID(ST_GeomFromText(%s), 26916)) "
                    "ON CONFLICT (geoid) DO UPDATE SET name=EXCLUDED.name, "
                    "geometry=EXCLUDED.geometry, load_run_id=EXCLUDED.load_run_id",
                    (
                        record.source_id,
                        record.name,
                        record.source_crs,
                        load_run_id,
                        geometry.wkt,
                    ),
                )
            elif record.dataset_id == "indygo_gtfs":
                connection.execute(
                    "INSERT INTO transit.stops(source_id, name, source_crs, "
                    "load_run_id, geometry) "
                    "VALUES (%s, %s, %s, %s, ST_SetSRID(ST_GeomFromText(%s), 26916)) "
                    "ON CONFLICT (source_id) DO UPDATE SET name=EXCLUDED.name, "
                    "geometry=EXCLUDED.geometry, load_run_id=EXCLUDED.load_run_id",
                    (
                        record.source_id,
                        record.name,
                        record.source_crs,
                        load_run_id,
                        geometry.wkt,
                    ),
                )
            elif record.service_type:
                connection.execute(
                    "INSERT INTO services.service_locations(source_id, service_type, "
                    "name, address, source_crs, load_run_id, geometry) VALUES "
                    "(%s, %s, %s, %s, %s, %s, ST_SetSRID(ST_GeomFromText(%s), 26916)) "
                    "ON CONFLICT (service_type, source_id) DO UPDATE SET "
                    "name=EXCLUDED.name, address=EXCLUDED.address, "
                    "geometry=EXCLUDED.geometry, load_run_id=EXCLUDED.load_run_id",
                    (
                        record.source_id,
                        record.service_type,
                        record.name,
                        record.address,
                        record.source_crs,
                        load_run_id,
                        geometry.wkt,
                    ),
                )
        for item in run.audit:
            connection.execute(
                "INSERT INTO etl.feature_audit(load_run_id, dataset_id, source_id, "
                "action, reason) "
                "VALUES (%s, %s, %s, %s, %s)",
                (
                    load_run_id,
                    item["dataset_id"],
                    item["source_id"],
                    item["action"],
                    item["reason"] or None,
                ),
            )
        connection.execute(
            "UPDATE etl.load_runs SET status='succeeded', completed_at=now() "
            "WHERE load_run_id=%s",
            (load_run_id,),
        )
    return run
